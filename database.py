"""
MIT License

Copyright (c) 2020 Daniel Silva dos Santos

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""


from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import collections.abc
import json
import pathlib


class TransPair(collections.UserList):
    """A translation pair is a list-like object that stores a sentence pair.

    A sentence pair must be a string object and empty string values are invalid.
    """
    # TODO: change initialization to accept *args for 2 item list,
    # and 2 string args
    def __init__(self, item0, item1):
        if not isinstance(item0, str) and not isinstance(item1, str):
            raise TypeError("string element expected.")
        elif (item0 and item1) == "":
            raise ValueError("cannot store empty strings.")
        super(TransPair, self).__init__([item0, item1])

    def __getattribute__(self, attr):
        """Prevent super class methods to break TransPair guides."""
        private_super = [
            "append", "extend", "insert",
            "pop", "remove", "clear", "sort"
        ]
        if attr in private_super:
            raise AttributeError
        else:
            return super(TransPair, self).__getattribute__(attr)

    def __setitem__(self, key, value):
        if not isinstance(value, str):
            raise TypeError("string element expected.")
        elif len(value) == 0:
            raise ValueError("cannot store empty strings.")
        super(TransPair, self).__setitem__(key, value)

    def switch(self):
        """Invert the TransPair order."""
        super(TransPair, self).reverse()


class TransList(collections.UserList):
    """A list of TransPair."""

    def __init__(self, *args):
        """
        :param args: TransPair or lists to construct TransPairs
        """
        super(TransList, self).__init__()
        for arg in args:
            self.append(arg)

    def __str__(self):
        indent = "    "
        trans_list_str = str()
        for i, item in enumerate(super(TransList, self).__iter__()):
            trans_list_str = trans_list_str + f"{indent} {str(item)}"
            if i < super(TransList, self).__len__() - 1:
                trans_list_str = trans_list_str + ", \n"
        return f"{{\n{trans_list_str}\n}}"

    def append(self, tp_key):
        """Adds a new TransPair element into the end of the TransList."""
        if not isinstance(tp_key, TransPair):
            raise TypeError("only TransPair object is accepted.")
        elif super(TransList, self).__contains__(tp_key):
            raise KeyError("TransPair {} already assigned.".format(tp_key))
        else:
            super(TransList, self).append(tp_key)

    def insert(self, i, tp_key):
        # TODO: test this method
        if type(tp_key) != TransPair:
            raise KeyError("the tp_key must be a TransPair object.")
        elif super(TransList, self).__contains__(tp_key):
            raise KeyError("TransPair {} already assigned.".format(tp_key))
        super(TransList, self).insert(i, tp_key)

    def extend(self, other):
        # TODO: test this method
        for trans_pair in other:
            self.append(trans_pair)

    def remove(self, tp_key):
        """Remove a TransPair element by element key."""
        super(TransList, self).remove(tp_key)

    def pop(self, index=-1):
        """Remove the item at the given position in the list, and return it.

        If no index is specified, a.pop() removes and returns the last item in
        the TransList.
        """
        super(TransList, self).pop(index)

    def get_translation(self, sentence, hint=0):
        """Return a the translation.

        Hint is where the search will take place. If 0 will search the sentence on
        all TransPairs elements. if 1 will do the search only on the first TransPairs
        index. If 2 will do the search only on the second TransPairs index.
        :param sentence:
        :param hint:
        :return: str
        """
        if hint == 0 or hint == 1:
            for transpair in self:
                if transpair[0] == sentence:
                    return transpair[1]
        if hint == 0 or hint == 2:
            for transpair in self:
                if transpair[1] == sentence:
                    return transpair[0]


class TransDatabase(dict):
    """Creates, load and manipulate TransLists.

    The TransDatabase are a dictionary that store and manipulate named TransList
    that manipulate TransPairs objects, and are constructed with some helpful
    attributes to organize your TransLists, load and manipulate TransPairs.
    In a TransDatabase, the TransList has a unique key to uniquely identify
    the TransList, but his content can be the same of other TransList.
    When you initialize a empty TransDatabase you need to identify a
    first_lang("e.g. EN-US") and second_lang("e.g. DE-CH") attribute. These are
    mere index for you to know the order in which you have to store the
    sentences, but not that does not stop the user store them wrong.
    """
    INFO_TITLE = "info"
    LANGS = "language"

    def __init__(self, first_lang=None, second_lang=None, **kwargs):
        super().__init__()
        if self.INFO_TITLE in kwargs:
            self.info = kwargs[self.INFO_TITLE]
            kwargs.pop(self.INFO_TITLE)
            self.info[self.LANGS] = self.__chk_lang_attr(first_lang, second_lang)
        else:
            if first_lang is None or second_lang is None:
                raise TypeError("first_lang or second_lang has no attribute")
            self.info = {
                self.LANGS: self.__chk_lang_attr(first_lang, second_lang),
                "creator": "Daniel Daninsky",
                "script": "duobreaker v0.1"
            }
        for tlist_name, decor_tlist in kwargs.items():
            tlist = TransList()
            for tpair in decor_tlist:
                tlist.append(TransPair(tpair[0], tpair[1]))
            self.add(tlist_name, tlist)

    def __setitem__(self, trans_list_name, trans_list):
        """Set self[key] to value.

        Caution: note that __setitem__ can overwrite data. So if you want add
        a new key with a overwrite garanty use the add method.
        """
        if not isinstance(trans_list_name, str):
            raise TypeError("invalid type")
        elif not isinstance(trans_list, TransList):
            TypeError("object isn't a TransList")
        else:
            super(TransDatabase, self).__setitem__(trans_list_name, trans_list)

    def add(self, trans_list_name, trans_list=None):
        """Add TransList with its name in the end of TransDatabase.

        If TransList is empty the effect are the same of the method add_trans_list.
        Add are recomended to add new TransList. Add has overwrite guard that
        __setitem__ doesn't.
        """
        if trans_list_name in self:
            raise KeyError("TransList already exists")
        if trans_list is None:
            trans_list = TransList()
        else:
            self.__setitem__(trans_list_name, trans_list)

    def get_translation(self, key, sentence, lang=None):
        """Get the translation."""
        if lang == self.info[self.LANGS][0]:
            return self[key].get_translation(sentence, 1)
        elif lang == self.info[self.LANGS][1]:
            return self[key].get_translation(sentence, 2)
        else:
            return self[key].get_translation(sentence, 0)

    def getbyindex(self, index):
        """Return TransList by index."""
        if index < 0 or index > len(self):
            raise IndexError("TransDatabase index out of range")
        for i, item in enumerate(self):
            if i == index:
                return self[item]

    def change_lang_attrs(self, first_lang, second_lang):
        """Change the language attributes."""
        self.info["language"] = TransPair(first_lang, second_lang)

    def save(self, file, overwrite=False):
        # TODO: implement overwrite guard
        if not isinstance(file, pathlib.Path):
            file = pathlib.Path(file)
        extension = file.suffix
        if extension == ".xlsx":
            self.__xlsx_save(file)
        elif extension == ".json":
            self.__json_save(file)
        else:
            raise ValueError("invalid extension type. xlsx or json")

    @classmethod
    def fromfile(cls, file):
        """Load a from disk a .xlsx translation database.

        :param file: Path-like object where the database will be opened
        :return: TransDatabase
        """
        if not isinstance(file, pathlib.Path):
            file = pathlib.Path(file)
        extension = file.suffix
        extension = extension.casefold()
        if extension == ".xlsx":
            kwargs = cls.__xlsx_load(file, cls.INFO_TITLE)
        elif extension == ".json":
            kwargs = cls.__json_load(file)
        else:
            raise ValueError("invalid extension file. xlsx or json")
        first_lang = kwargs[cls.INFO_TITLE][cls.LANGS][0]  # confusing AF
        second_lang = kwargs[cls.INFO_TITLE][cls.LANGS][1]
        return cls(first_lang, second_lang, **kwargs)

    def __xlsx_save(self, file):
        # TODO: improve style and add a info style
        workbook = Workbook()
        worksheet = workbook.active  # current working spreadsheet
        workbook.remove(worksheet)

        def sheet_decor(ws):
            """Sheet style fo here, font, size, etc."""
            default_font = Font(name='Arial', size=12)
            ws.column_dimensions['A'].font = default_font
            ws.column_dimensions['B'].font = default_font
            cell_width = 60
            ws.column_dimensions['A'].width = cell_width
            ws.column_dimensions['B'].width = cell_width

        # add content
        for key, value in self.items():
            worksheet = workbook.create_sheet(key)
            sheet_decor(worksheet)
            for num, trans_pair in enumerate(value, start=1):
                worksheet.cell(row=num, column=1).value = trans_pair[0]
                worksheet.cell(row=num, column=2).value = trans_pair[1]
        # add info
        worksheet_info = workbook.create_sheet(self.INFO_TITLE)
        worksheet_info["A1"] = json.dumps(self.info)
        workbook.save(file)

    def __json_save(self, file):
        std_dict = {}
        for key, value in self.items():
            std_list = []
            for trans_pair in value:
                std_list.append(list(trans_pair))
            std_dict[key] = std_list
        std_dict[self.INFO_TITLE] = self.info
        file_obj = open(file, "w")
        json.dump(std_dict, file_obj, indent="    ", ensure_ascii=False)
        file_obj.close()

    @staticmethod
    def __xlsx_load(file, info_title):

        workbook = load_workbook(file)
        kwargs = {}
        if info_title in workbook.sheetnames:
            kwargs[info_title] = json.loads(workbook[info_title]["A1"].value)
            workbook.remove(workbook[info_title])
        for spreadsheet in workbook.worksheets:
            decor_tlist = []
            spreadsheet_name = spreadsheet.title
            for row in spreadsheet.iter_rows(min_row=1, max_col=2, values_only=True):
                # TODO: empty cell guard needs improviment
                if row[0] is None or row[1] is None:
                    break
                decor_tlist.append([row[0], row[1]])
            kwargs[spreadsheet_name] = decor_tlist
        return kwargs

    @staticmethod
    def __json_load(file):
        file = open(file, "r")
        kwargs = json.load(file)
        file.close()
        return kwargs

    @staticmethod
    def __chk_lang_attr(fl, sl):
        if not isinstance(fl, str):
            fl = str(fl)
        if not isinstance(sl, str):
            sl = str(sl)
        if fl == sl:
            raise ValueError("languages has the same attribute value.")
        return [fl, sl]


if __name__ == '__main__':
    pass
